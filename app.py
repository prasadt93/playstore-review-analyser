# ═══════════════════════════════════════════════════════════════════════════════
#  Play Store Review Analyser  ·  Streamlit App  v3
#  No API keys · No AI · Runs fully offline on your Mac
# ═══════════════════════════════════════════════════════════════════════════════

import streamlit as st
import json, os, re, io, warnings
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")

from google_play_scraper import reviews as gps_reviews, app as gps_app, Sort
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "review_history.json")
MONTHS_OPTIONS = {"Last 1 Month":1,"Last 3 Months":3,"Last 6 Months":6,"Last 12 Months":12}

sia = SentimentIntensityAnalyzer()

# ══════════════════════════════════════════════════════════════════════════════
#  HISTORY
# ══════════════════════════════════════════════════════════════════════════════

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE,"r") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_history(history):
    with open(HISTORY_FILE,"w") as f:
        json.dump(history,f,indent=2)

def add_to_history(app_id,app_name,url,icon_url=""):
    history=load_history()
    history=[h for h in history if h.get("app_id")!=app_id]
    history.insert(0,{"app_id":app_id,"app_name":app_name,"url":url,
                       "icon_url":icon_url,"last_searched":datetime.now().isoformat()})
    save_history(history[:25])

# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPING
# ══════════════════════════════════════════════════════════════════════════════

def extract_app_id(url):
    m=re.search(r"[?&]id=([a-zA-Z0-9_.]+)",url)
    return m.group(1) if m else None

@st.cache_data(ttl=3600,show_spinner=False)
def fetch_app_info(app_id):
    for country in ["in","us"]:
        try:
            return gps_app(app_id,lang="en",country=country)
        except Exception:
            continue
    return {}

@st.cache_data(ttl=3600,show_spinner=False)
def fetch_reviews_cached(app_id,months):
    cutoff=datetime.now()-timedelta(days=months*30.5)
    all_reviews,token=[],None
    for _ in range(60):
        try:
            batch,token=gps_reviews(app_id,lang="en",country="in",
                                     sort=Sort.NEWEST,count=200,continuation_token=token)
        except Exception:
            break
        if not batch:
            break
        for r in batch:
            rev_date=r.get("at")
            if isinstance(rev_date,datetime) and rev_date<cutoff:
                return all_reviews
            all_reviews.append(r)
        if not token:
            break
    return all_reviews

# ══════════════════════════════════════════════════════════════════════════════
#  ANALYSIS ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def sent_label(text):
    s=sia.polarity_scores(str(text))["compound"]
    return "Positive" if s>=0.05 else ("Negative" if s<=-0.05 else "Neutral")

def word_count(text):
    return len(str(text).split())

def analyse_reviews(df):
    df=df.copy()
    df["at"]=pd.to_datetime(df["at"])
    df["sentiment"]=df["content"].apply(sent_label)
    df["compound"] =df["content"].apply(lambda x:sia.polarity_scores(str(x))["compound"])
    df["word_count"]=df["content"].apply(word_count)

    results={"df":df,"total_reviews":len(df),
             "avg_rating":round(df["score"].mean(),2),
             "rating_dist":df["score"].value_counts().sort_index(),
             "sentiment_dist":df["sentiment"].value_counts()}

    replied=df["replyContent"].notna().sum() if "replyContent" in df.columns else 0
    results["reply_rate"]=round(replied/max(len(df),1)*100,1)

    df["month"]=df["at"].dt.to_period("M")
    monthly=df.groupby("month").agg(
        count=("score","count"),avg_rating=("score","mean"),
        one_star=("score",lambda x:(x==1).sum()),
        two_star=("score",lambda x:(x==2).sum()),
        three_star=("score",lambda x:(x==3).sum()),
        four_star=("score",lambda x:(x==4).sum()),
        five_star=("score",lambda x:(x==5).sum()),
        positive=("sentiment",lambda x:(x=="Positive").sum()),
        negative=("sentiment",lambda x:(x=="Negative").sum()),
        neutral=("sentiment", lambda x:(x=="Neutral").sum()),
    ).reset_index()
    monthly["month_str"]=monthly["month"].astype(str)
    monthly["avg_rating"]=monthly["avg_rating"].round(2)
    monthly["neg_pct"]=((monthly["one_star"]+monthly["two_star"])/monthly["count"]*100).round(1)
    monthly["pos_pct"]=((monthly["four_star"]+monthly["five_star"])/monthly["count"]*100).round(1)
    results["monthly"]=monthly

    last30=df[df["at"]>=datetime.now()-timedelta(days=30)]
    results["recent_avg"]=round(last30["score"].mean(),2) if len(last30) else results["avg_rating"]

    return results

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR  — 2-sheet report: Dashboard + All Reviews
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel(analysis,app_name,months):
    wb=Workbook()
    monthly=analysis["monthly"]
    month_cols=monthly["month_str"].tolist()

    # ── Style constants ───────────────────────────────────────────────────────
    DB="1F4E79"; MB="2E75B6"; LB="D6E4F0"
    db_f=PatternFill("solid",fgColor=DB); mb_f=PatternFill("solid",fgColor=MB)
    lb_f=PatternFill("solid",fgColor=LB); alt_f=PatternFill("solid",fgColor="EBF3FB")
    wh_f=PatternFill("solid",fgColor="FFFFFF")
    H1=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
    H2=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    H3=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
    BD=Font(name="Calibri",size=10); BD9=Font(name="Calibri",size=9)
    CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LFT=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    thin=Side(style="thin",color="BFBFBF")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

    def hdr_row(ws,ri,fill,fnt=H3):
        for c in ws[ri]:
            c.fill=fill; c.font=fnt; c.alignment=CTR; c.border=bdr

    def safe_int(x):
        try: return int(x) if x is not None and not (isinstance(x,float) and np.isnan(x)) else 0
        except: return 0

    # ══════════════════════════════════════════════════════════════════════════
    # S1 : Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    ws1=wb.active; ws1.title="📊 Dashboard"
    ws1.column_dimensions["A"].width=14
    for col in list("BCDEFGHIJKLMNO"):
        ws1.column_dimensions[col].width=12

    ws1.merge_cells("A1:N1")
    ws1["A1"]=f"{app_name} — PlayStore Review Analysis"
    ws1["A1"].font=H1; ws1["A1"].fill=db_f; ws1["A1"].alignment=CTR; ws1.row_dimensions[1].height=24

    ws1.merge_cells("A2:N2")
    ws1["A2"]=(f"Period: {month_cols[0]} – {month_cols[-1]}  |  Total Reviews: {analysis['total_reviews']:,}"
               f"  |  Avg Rating: {analysis['avg_rating']} ★  |  Generated: {datetime.now().strftime('%d %b %Y')}")
    ws1["A2"].font=Font(name="Calibri",size=10,color="FFFFFF")
    ws1["A2"].fill=PatternFill("solid",fgColor="2E4057"); ws1["A2"].alignment=CTR; ws1.row_dimensions[2].height=18

    # KPI tiles
    kpi_data=[
        ("Peak Month",f"{monthly.loc[monthly['count'].idxmax(),'month_str']} ({monthly['count'].max():,})","Highest review volume","2E75B6"),
        ("Avg Rating",f"{analysis['avg_rating']} / 5.0","Across all months","375623"),
        ("Best Rating Month",f"{monthly.loc[monthly['avg_rating'].idxmax(),'month_str']} ({monthly['avg_rating'].max():.2f}★)","Highest avg rating","7030A0"),
        ("Lowest Neg% Month",f"{monthly.loc[monthly['neg_pct'].idxmin(),'month_str']} ({monthly['neg_pct'].min():.1f}%)","Fewest 1+2★ reviews","C00000"),
    ]
    kpi_col_pairs=[("B","D"),("E","G"),("H","J"),("K","M")]
    for (sc,ec),(title,val,sub,color) in zip(kpi_col_pairs,kpi_data):
        f=PatternFill("solid",fgColor=color)
        for row,txt,fnt in [(4,title,Font(name="Calibri",bold=True,size=9,color="FFFFFF")),
                             (5,val,  Font(name="Calibri",bold=True,size=13,color="FFFFFF")),
                             (6,sub,  Font(name="Calibri",size=9,color=LB))]:
            ws1.merge_cells(f"{sc}{row}:{ec}{row}")
            ws1[f"{sc}{row}"]=txt; ws1[f"{sc}{row}"].fill=f
            ws1[f"{sc}{row}"].font=fnt; ws1[f"{sc}{row}"].alignment=CTR
        ws1.row_dimensions[5].height=26

    # Monthly table
    ws1.merge_cells("A8:J8")
    ws1["A8"]="Monthly Rating Summary"
    ws1["A8"].font=H2; ws1["A8"].fill=db_f; ws1["A8"].alignment=CTR; ws1.row_dimensions[8].height=18

    mhdrs=["Month","Total Reviews","★ Avg","1★","2★","3★","4★","5★","1+2★ %","4+5★ %"]
    for ci,h in enumerate(mhdrs,1):
        ws1.cell(9,ci,h).fill=mb_f; ws1.cell(9,ci).font=H3
        ws1.cell(9,ci).alignment=CTR; ws1.cell(9,ci).border=bdr

    for ri,(_, row) in enumerate(monthly.iterrows(),start=10):
        fill=alt_f if ri%2==0 else wh_f
        for ci,v in enumerate([row["month_str"],safe_int(row["count"]),row["avg_rating"],
                                safe_int(row["one_star"]),safe_int(row["two_star"]),safe_int(row["three_star"]),
                                safe_int(row["four_star"]),safe_int(row["five_star"]),
                                f'{row["neg_pct"]}%',f'{row["pos_pct"]}%'],1):
            c=ws1.cell(ri,ci,v); c.fill=fill; c.font=BD9; c.alignment=CTR; c.border=bdr
        ws1.row_dimensions[ri].height=15

    # ══════════════════════════════════════════════════════════════════════════
    # S2 : All Reviews
    # ══════════════════════════════════════════════════════════════════════════
    ws2=wb.create_sheet("📋 All Reviews")
    df_all=analysis["df"].copy()

    SENT_FILLS={"Positive":PatternFill("solid",fgColor="E2EFDA"),
                "Negative":PatternFill("solid",fgColor="FCE4D6"),
                "Neutral": PatternFill("solid",fgColor="F5F5F5")}
    WC_FILL=PatternFill("solid",fgColor="EBF3FB")   # light blue for word count column

    base_cols={"userName":"Reviewer","score":"Rating","content":"Review Text",
               "at":"Date","thumbsUpCount":"Helpful","replyContent":"Dev Reply",
               "sentiment":"Sentiment","word_count":"Word Count"}
    avail=[c for c in base_cols if c in df_all.columns]
    headers=[base_cols[c] for c in avail]

    ws2.append(headers)
    hdr_row(ws2,1,mb_f,H3)

    cw={"Reviewer":18,"Rating":8,"Review Text":65,"Date":12,
        "Helpful":10,"Dev Reply":40,"Sentiment":12,"Word Count":12}
    for i,h in enumerate(headers,1):
        ws2.column_dimensions[get_column_letter(i)].width=cw.get(h,15)

    # find word_count column index (1-based)
    wc_col_idx = avail.index("word_count") + 1 if "word_count" in avail else None

    disp=df_all[avail].copy()
    disp["at"]=pd.to_datetime(disp["at"]).dt.strftime("%Y-%m-%d")

    for idx,rv in enumerate(disp.itertuples(index=False)):
        row_vals=list(rv)
        ws2.append(row_vals)
        ri=ws2.max_row
        sent=rv._asdict().get("sentiment","Neutral")
        base_fill=SENT_FILLS.get(sent,SENT_FILLS["Neutral"])
        for ci,cell in enumerate(ws2[ri],1):
            if wc_col_idx and ci == wc_col_idx:
                cell.fill=WC_FILL
                cell.font=Font(name="Calibri",size=9,bold=True,color="1F4E79")
                cell.alignment=Alignment(horizontal="center",vertical="center")
            else:
                cell.fill=base_fill; cell.font=BD9
                cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
            cell.border=bdr
        ws2.row_dimensions[ri].height=35

    ws2.freeze_panes="A2"; ws2.auto_filter.ref=ws2.dimensions

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Play Store Review Analyser",page_icon="📱",
                   layout="wide",initial_sidebar_state="expanded")

st.markdown("""
<style>
[data-testid="stSidebar"]{background:linear-gradient(180deg,#1F4E79 0%,#154069 100%);}
[data-testid="stSidebar"] p,[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,[data-testid="stSidebar"] label{color:#E8F1FB !important;}
.stButton>button{background:linear-gradient(135deg,#1F4E79,#2E75B6);color:white;border:none;
  border-radius:8px;padding:8px 18px;font-weight:600;transition:all 0.2s ease;}
.stButton>button:hover{background:linear-gradient(135deg,#2E75B6,#1F4E79);
  box-shadow:0 4px 12px rgba(31,78,121,0.35);transform:translateY(-1px);}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#375623,#70AD47) !important;
  color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important;}
[data-testid="stMetric"]{background:linear-gradient(135deg,#EBF3FB,#D6E4F0);border-radius:10px;
  padding:12px 16px;border-left:4px solid #2E75B6;box-shadow:0 1px 4px rgba(0,0,0,0.08);}
[data-testid="stSidebar"] .stButton>button{background:rgba(255,255,255,0.12) !important;
  border:1px solid rgba(255,255,255,0.25) !important;border-radius:8px !important;
  text-align:left !important;padding:6px 10px !important;font-size:13px !important;
  line-height:1.4 !important;color:#E8F1FB !important;margin-bottom:2px;}
[data-testid="stSidebar"] .stButton>button:hover{background:rgba(255,255,255,0.22) !important;transform:none;}
</style>
""",unsafe_allow_html=True)

for k in ["prefill_url","results","excel_cache","cache_key"]:
    if k not in st.session_state:
        st.session_state[k]=None

with st.sidebar:
    st.markdown("## 📱 Review Analyser"); st.markdown("---")
    if st.button("➕  New Search",use_container_width=True):
        for k in ["prefill_url","results","excel_cache","cache_key"]:
            st.session_state[k]=None
        st.rerun()
    st.markdown("### 🕑 History")
    history=load_history()
    if not history:
        st.caption("No history yet.")
    else:
        for item in history:
            c1,c2=st.columns([5,1])
            with c1:
                if st.button(f"{item['app_name'][:24]}\n🗓 {item['last_searched'][:10]}",
                             key=f"h_{item['app_id']}",use_container_width=True):
                    st.session_state.prefill_url=item["url"]
                    st.session_state.results=None; st.session_state.excel_cache=None; st.rerun()
            with c2:
                if st.button("✕",key=f"d_{item['app_id']}",help="Remove"):
                    history=[h for h in history if h["app_id"]!=item["app_id"]]
                    save_history(history); st.rerun()

st.markdown('<h1 style="color:#1F4E79;margin-bottom:2px">📱 Play Store Review Analyser</h1>',unsafe_allow_html=True)
st.caption("Paste any Play Store URL · choose a time period · get charts and a full raw-data Excel report.")
st.markdown("---")

c_url,c_dur,c_btn=st.columns([4,2,1])
with c_url:
    url_input=st.text_input("🔗 Play Store App URL",value=st.session_state.prefill_url or "",
                             placeholder="https://play.google.com/store/apps/details?id=com.example.app")
with c_dur:
    dur_label=st.selectbox("📅 Time Period",list(MONTHS_OPTIONS.keys()),index=3)
    months=MONTHS_OPTIONS[dur_label]
with c_btn:
    st.markdown("<br>",unsafe_allow_html=True)
    get_data=st.button("🔍  Get Data",type="primary",use_container_width=True)

if get_data:
    if not url_input.strip():
        st.warning("Please enter a Play Store URL.")
    else:
        app_id=extract_app_id(url_input.strip())
        if not app_id:
            st.error("⚠️ Couldn't find an app ID in the URL.")
        else:
            with st.spinner(f"Fetching & analysing reviews for the last {months} month(s)…"):
                try:
                    app_info=fetch_app_info(app_id)
                    raw=fetch_reviews_cached(app_id,months)
                    if not raw:
                        st.warning("No reviews found for this period.")
                    else:
                        df=pd.DataFrame(raw)
                        analysis=analyse_reviews(df)
                        app_name=app_info.get("title",app_id)
                        add_to_history(app_id,app_name,url_input.strip(),app_info.get("icon",""))
                        st.session_state.results=dict(analysis=analysis,
                            app_name=app_name,app_info=app_info,months=months,
                            dur_label=dur_label,app_id=app_id)
                        st.session_state.excel_cache=None
                        st.session_state.cache_key=f"{app_id}_{months}"
                        st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")

if st.session_state.results:
    res=st.session_state.results
    analysis=res["analysis"]
    app_name=res["app_name"]; app_info=res["app_info"]
    months=res["months"]; app_id=res["app_id"]

    ic,inf=st.columns([1,9])
    with ic:
        if app_info.get("icon"): st.image(app_info["icon"],width=60)
    with inf:
        st.markdown(f"### {app_name}")
        st.caption(f"**Developer:** {app_info.get('developer','—')}  ·  "
                   f"**Category:** {app_info.get('genre','—')}  ·  "
                   f"**Overall Rating:** {app_info.get('score','—')} ⭐  ·  Period: {res['dur_label']}")
    st.markdown("---")

    total=analysis["total_reviews"]
    pos=int(analysis["sentiment_dist"].get("Positive",0))
    neg=int(analysis["sentiment_dist"].get("Negative",0))
    k1,k2,k3,k4,k5,k6=st.columns(6)
    k1.metric("📊 Total Reviews",f"{total:,}")
    k2.metric("⭐ Avg Rating",analysis["avg_rating"])
    k3.metric("📅 Recent 30-Day",analysis["recent_avg"])
    k4.metric("😊 Positive",f"{pos} ({round(pos/max(total,1)*100,1)}%)")
    k5.metric("😞 Negative",f"{neg} ({round(neg/max(total,1)*100,1)}%)")
    k6.metric("💬 Dev Reply Rate",f"{analysis['reply_rate']}%")

    st.markdown("<br>",unsafe_allow_html=True)

    tab1,tab2=st.tabs(["📈 Charts","📋 Raw Data"])

    with tab1:
        cl,cr=st.columns(2)
        with cl:
            rd=analysis["rating_dist"].reset_index(); rd.columns=["Stars","Count"]
            rd["Stars"]=rd["Stars"].astype(str)+" ⭐"
            fig=px.bar(rd,x="Stars",y="Count",color="Count",text="Count",
                       color_continuous_scale=["#C00000","#ED7D31","#FFC000","#70AD47","#375623"],
                       title="Rating Distribution")
            fig.update_traces(textposition="outside")
            fig.update_layout(showlegend=False,coloraxis_showscale=False,
                              plot_bgcolor="white",paper_bgcolor="white",margin=dict(t=40,b=10,l=10,r=10))
            st.plotly_chart(fig,use_container_width=True)
        with cr:
            sd=analysis["sentiment_dist"].reset_index(); sd.columns=["Sentiment","Count"]
            fig2=px.pie(sd,values="Count",names="Sentiment",hole=0.42,title="Sentiment Distribution",
                        color="Sentiment",color_discrete_map={"Positive":"#375623","Negative":"#C00000","Neutral":"#FFC000"})
            fig2.update_layout(plot_bgcolor="white",paper_bgcolor="white",margin=dict(t=40,b=10,l=10,r=10))
            st.plotly_chart(fig2,use_container_width=True)
        if len(analysis["monthly"])>1:
            m=analysis["monthly"]
            fig3=go.Figure()
            fig3.add_trace(go.Bar(name="Positive",x=m["month_str"],y=m["positive"],marker_color="#375623"))
            fig3.add_trace(go.Bar(name="Neutral", x=m["month_str"],y=m["neutral"], marker_color="#FFC000"))
            fig3.add_trace(go.Bar(name="Negative",x=m["month_str"],y=m["negative"],marker_color="#C00000"))
            fig3.add_trace(go.Scatter(name="Avg Rating",x=m["month_str"],y=m["avg_rating"],
                                       yaxis="y2",mode="lines+markers",
                                       line=dict(color="#1F4E79",width=2.5),marker=dict(size=6)))
            fig3.update_layout(barmode="stack",title="Monthly Volume & Sentiment",
                                yaxis=dict(title="Review Count"),
                                yaxis2=dict(title="Avg Rating",overlaying="y",side="right",range=[0,5.5]),
                                plot_bgcolor="white",paper_bgcolor="white",
                                legend=dict(orientation="h",yanchor="bottom",y=1.02),
                                margin=dict(t=60,b=10,l=10,r=10))
            st.plotly_chart(fig3,use_container_width=True)

    with tab2:
        df_disp=analysis["df"].copy()
        cols=["userName","score","sentiment","word_count","content","at","thumbsUpCount"]
        df_disp=df_disp[[c for c in cols if c in df_disp.columns]].copy()
        df_disp["at"]=pd.to_datetime(df_disp["at"]).dt.strftime("%Y-%m-%d")
        df_disp.columns=[c.replace("userName","Reviewer").replace("score","Rating")
                          .replace("sentiment","Sentiment").replace("word_count","Word Count")
                          .replace("content","Review").replace("at","Date")
                          .replace("thumbsUpCount","Helpful") for c in df_disp.columns]
        st.dataframe(df_disp,use_container_width=True,height=520)

    st.markdown("---")
    dl1,_=st.columns([2,6])
    cache_key=f"{app_id}_{months}"
    if st.session_state.cache_key!=cache_key or st.session_state.excel_cache is None:
        with st.spinner("Building Excel report…"):
            st.session_state.excel_cache=generate_excel(analysis,app_name,months)
        st.session_state.cache_key=cache_key
    with dl1:
        st.download_button("📥 Download Excel Report",
                           data=st.session_state.excel_cache,
                           file_name=f"{app_id}_{months}mo_reviews.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
